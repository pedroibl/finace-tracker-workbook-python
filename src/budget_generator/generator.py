"""Workbook generator responsible for orchestrating sheet creation."""

from __future__ import annotations

import logging
from collections.abc import Iterable
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .charts import add_dashboard_doughnut_charts
from .sheets.calculations import (
    build_calculations_sheet,
    register_calculations_named_ranges,
)
from .sheets.dashboard import build_dashboard_sheet, register_dashboard_named_ranges
from .sheets.dropdown import build_dropdown_sheet, register_dropdown_named_ranges
from .sheets.planning import build_planning_sheet, register_planning_named_ranges
from .sheets.settings import build_settings_sheet, register_settings_named_ranges
from .sheets.tracking import build_tracking_sheet
from .utils.named_ranges import NamedRangeManager


class GeneratorError(RuntimeError):
    """Base error for generator failures."""


class WorkbookNotInitialisedError(GeneratorError):
    """Raised when sheet operations occur before a workbook exists."""


class SheetMissingError(GeneratorError):
    """Raised when the expected sheet is absent from the workbook."""


class BudgetGenerator:
    """Generate the Excel workbook defined by the specification."""

    def __init__(self, spec: Mapping[str, Any]):
        self.spec = spec
        self.workbook: Workbook | None = None

    # ------------------------------------------------------------------
    # Workbook structure helpers
    # ------------------------------------------------------------------
    def create_workbook(self) -> Workbook:
        """Create a new workbook and remove the default sheet."""

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        self.workbook = workbook
        return workbook

    def create_sheets(self, spec: Mapping[str, Any] | None = None) -> None:
        """Instantiate worksheets using the workbook.sheets metadata."""

        workbook = self._require_workbook()
        active_spec = spec or self.spec
        workbook_spec = active_spec.get("workbook", {})
        sheets_spec: Iterable[Mapping[str, Any]] = workbook_spec.get("sheets", [])

        for sheet_meta in sheets_spec:
            name = sheet_meta.get("name")
            visibility = sheet_meta.get("visibility", "visible")
            if not isinstance(name, str) or not name:
                raise GeneratorError("Sheet metadata must include a non-empty 'name'.")

            worksheet = workbook.create_sheet(title=name)
            if visibility in {"hidden", "veryHidden"}:
                worksheet.sheet_state = visibility

    def build_sheet_contents(self) -> None:
        """Populate worksheets and register named ranges according to the PRD."""

        workbook = self._require_workbook()
        sheet_specs = self._sheet_specs()
        manager = NamedRangeManager(workbook)

        settings_ws = self._get_sheet("Settings")
        LOGGER.info("Building Settings sheet")
        build_settings_sheet(settings_ws, sheet_specs.get("Settings", {}))
        register_settings_named_ranges(manager)

        dropdown_ws = self._get_sheet("Dropdown Data")
        LOGGER.info("Building Dropdown Data sheet")
        build_dropdown_sheet(dropdown_ws, sheet_specs.get("Dropdown Data", {}))
        register_dropdown_named_ranges(manager)

        planning_ws = self._get_sheet("Budget-Planning")
        LOGGER.info("Building Budget-Planning sheet")
        build_planning_sheet(planning_ws, sheet_specs.get("Budget-Planning", {}))
        register_planning_named_ranges(manager)

        tracking_ws = self._get_sheet("Budget Tracking")
        LOGGER.info("Building Budget Tracking sheet")
        build_tracking_sheet(tracking_ws, sheet_specs.get("Budget Tracking", {}))

        calculations_ws = self._get_sheet("Calculations")
        LOGGER.info("Building Calculations sheet")
        build_calculations_sheet(calculations_ws, sheet_specs.get("Calculations", {}))
        register_calculations_named_ranges(manager)

        dashboard_ws = self._get_sheet("Budget Dashboard")
        LOGGER.info("Building Dashboard sheet")
        build_dashboard_sheet(dashboard_ws, sheet_specs.get("Budget Dashboard", {}))
        register_dashboard_named_ranges(manager)
        add_dashboard_doughnut_charts(dashboard_ws)

        # Ensure helper sheets remain hidden.
        for sheet_name in ("Dropdown Data", "Calculations"):
            if sheet_name in workbook.sheetnames:
                workbook[sheet_name].sheet_state = "hidden"

    def save_workbook(self, output_path: Path) -> Path:
        """Persist the workbook to disk while surfacing I/O errors clearly."""

        workbook = self._require_workbook()

        output_path = Path(output_path)
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output_path)
        except OSError as exc:  # pragma: no cover - relies on OS failures
            raise GeneratorError(f"Failed to write workbook to {output_path}: {exc}") from exc

        return output_path

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------
    def _require_workbook(self) -> Workbook:
        if self.workbook is None:
            raise WorkbookNotInitialisedError("Call create_workbook() before using the workbook.")
        return self.workbook

    def _sheet_specs(self) -> dict[str, Mapping[str, Any]]:
        sheets_config = self.spec.get("sheets", {})
        if not isinstance(sheets_config, Mapping):
            return {}
        result: dict[str, Mapping[str, Any]] = {}
        for name, cfg in sheets_config.items():
            if isinstance(cfg, Mapping):
                result[str(name)] = cfg
        return result

    def _get_sheet(self, name: str) -> Worksheet:
        workbook = self._require_workbook()
        try:
            return workbook[name]
        except KeyError as exc:  # pragma: no cover - defensive guard
            raise SheetMissingError(f"Expected worksheet '{name}' to exist") from exc


LOGGER = logging.getLogger(__name__)
