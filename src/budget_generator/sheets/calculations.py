"""Calculations worksheet builder for metrics and budget comparisons."""

from __future__ import annotations

from typing import Mapping, Sequence

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from ..formulas.calculations import (
    build_choose_month_formula,
    build_monthly_tracking_sumproduct,
)
from ..utils.named_ranges import NamedRangeManager, NamedRangeSpec
from .planning import ACCOUNTING_FORMAT, MONTHS

METRIC_HEADER_FILL = "EAD1DC"
METRIC_HEADER_VALUES = ("Metric", "Value", "Notes")


class CalculationsSheetBuilder:
    """Encapsulates Calculations sheet generation logic."""

    def __init__(self, worksheet: Worksheet, spec: Mapping[str, object] | None = None):
        self.ws = worksheet
        self.spec = spec or {}

    def build(self) -> None:
        self._build_metric_tiles()
        self._build_month_map()
        self._build_budget_vs_tracked_table()

    def _build_metric_tiles(self) -> None:
        header_fill = PatternFill(start_color=METRIC_HEADER_FILL, end_color=METRIC_HEADER_FILL, fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")

        for column, value in enumerate(METRIC_HEADER_VALUES, start=2):
            cell = self.ws.cell(row=2, column=column, value=value)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

        metrics: Sequence[tuple[str, str, str]] = (
            ("Current Date", "=TODAY()", ""),
            ("Last Record Date", "=MAX(tblTracking[Date])", ""),
            ("Number of Records", "=COUNTA(tblTracking[Date])", ""),
            (
                "Tracking Balance",
                "=IFERROR(LOOKUP(2,1/(tblTracking[Date]<>\"\"),tblTracking[Balance]),0)",
                "",
            ),
        )

        for index, (label, formula, notes) in enumerate(metrics, start=3):
            label_cell = self.ws.cell(row=index, column=2, value=label)
            if index == 6:
                label_cell.font = Font(bold=True)
            value_cell = self.ws.cell(row=index, column=3, value=formula)
            notes_cell = self.ws.cell(row=index, column=4, value=notes)

            if index in {3, 4}:
                value_cell.number_format = "yyyy-mm-dd"
            elif index == 6:
                value_cell.number_format = ACCOUNTING_FORMAT

            notes_cell.alignment = Alignment(wrap_text=True)

        self._apply_border("B2", "D6")

    def _build_month_map(self) -> None:
        month_column = 10  # Column J
        index_column = 11  # Column K

        for offset, month in enumerate(MONTHS):
            row = 2 + offset
            self.ws.cell(row=row, column=month_column, value=month)
            self.ws.cell(row=row, column=index_column, value=offset + 1)

        month_idx_cell = self.ws.cell(row=1, column=index_column)
        month_idx_cell.value = "=INDEX(INDEX(MonthMap,0,2),MATCH(DashPeriod,INDEX(MonthMap,0,1),0))"

    def _build_budget_vs_tracked_table(self) -> None:
        header_fill = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")

        headers = ("Section", "BudgetedMonth", "TrackedMonth", "Remaining")
        for column_offset, title in enumerate(headers, start=5):
            cell = self.ws.cell(row=2, column=column_offset, value=title)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

        budget_rows = {
            3: ("Income", 13, "Income"),
            4: ("Expenses", 26, "Expense"),
            5: ("Savings", 34, "Saving"),
        }

        month_columns = [get_column_letter(index) for index in range(4, 16)]

        for row, (label, planning_row, tracking_type) in budget_rows.items():
            self.ws.cell(row=row, column=5, value=label)

            choose_cells = [f"{column_letter}{planning_row}" for column_letter in month_columns]
            choose_formula = build_choose_month_formula(choose_cells)
            tracked_formula = build_monthly_tracking_sumproduct(tracking_type)

            budget_cell = self.ws.cell(row=row, column=6, value=choose_formula)
            tracked_cell = self.ws.cell(row=row, column=7, value=tracked_formula)
            remaining_cell = self.ws.cell(row=row, column=8, value=f"=F{row}-G{row}")

            budget_cell.number_format = ACCOUNTING_FORMAT
            tracked_cell.number_format = ACCOUNTING_FORMAT
            remaining_cell.number_format = ACCOUNTING_FORMAT

        self._apply_border("E2", "H5")

    def _apply_border(self, start_cell: str, end_cell: str) -> None:
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        start_column = column_index_from_string(self.ws[start_cell].column_letter)
        start_row = self.ws[start_cell].row
        end_column = column_index_from_string(self.ws[end_cell].column_letter)
        end_row = self.ws[end_cell].row

        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                self.ws.cell(row=row, column=column).border = border


def build_calculations_sheet(worksheet: Worksheet, spec: Mapping[str, object] | None = None) -> None:
    CalculationsSheetBuilder(worksheet, spec).build()


def register_calculations_named_ranges(manager: NamedRangeManager) -> None:
    specs = (
        NamedRangeSpec("MonthMap", "Calculations", "$J$2:$K$13"),
        NamedRangeSpec("MonthIdx", "Calculations", "$K$1"),
    )
    manager.register_many(specs)
