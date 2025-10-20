"""Budget Dashboard worksheet builder."""

from __future__ import annotations

from typing import Mapping

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from ..utils.named_ranges import NamedRangeManager, NamedRangeSpec

HEADER_FILL = "DAEEF3"
SELECTOR_FILL = "E7F3F9"
TILE_FILL = "F9FBFD"
HEADER_VALUES = (
    "Year",
    "Period",
    "Category",
    "Tracked",
    "Budgeted",
    "% of Budget",
    "Remaining",
)


def build_dashboard_sheet(worksheet: Worksheet, spec: Mapping[str, object] | None = None) -> None:
    """Build the dashboard structure with selectors and table headers."""

    config = spec or {}
    selectors_config = (
        config.get("selectors", {}) if isinstance(config, Mapping) else {}
    )
    tiles_config = config.get("tiles", {}) if isinstance(config, Mapping) else {}
    default_period = selectors_config.get("default_period", "Jan")
    default_year_formula = selectors_config.get("default_year_formula", "=StartingYear")
    tracking_balance_formula = tiles_config.get("tracking_balance_formula", "=Calculations!C6")
    savings_rate_formula = tiles_config.get(
        "savings_rate_formula",
        "=IFERROR(Calculations!G5/SUM(Calculations!F3:F5),0)",
    )

    _build_header_row(worksheet)
    _build_selectors(worksheet, default_year_formula, default_period)
    _build_kpi_tiles(worksheet, tracking_balance_formula, savings_rate_formula)


def _build_header_row(worksheet: Worksheet) -> None:
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    header_fill = PatternFill(
        start_color=HEADER_FILL,
        end_color=HEADER_FILL,
        fill_type="solid",
    )

    for column_index, value in enumerate(HEADER_VALUES, start=2):  # column B onwards
        cell = worksheet.cell(row=2, column=column_index, value=value)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill


def _build_selectors(
    worksheet: Worksheet,
    default_year_formula: str,
    default_period: str,
) -> None:
    label_font = Font(bold=True)
    selector_alignment = Alignment(horizontal="center")
    selector_fill = PatternFill(
        start_color=SELECTOR_FILL,
        end_color=SELECTOR_FILL,
        fill_type="solid",
    )

    worksheet["B3"].value = "Year"
    worksheet["B3"].font = label_font

    worksheet["B4"].value = "Period"
    worksheet["B4"].font = label_font

    year_cell = worksheet["C3"]
    year_cell.value = default_year_formula
    year_cell.alignment = selector_alignment
    year_cell.fill = selector_fill

    period_cell = worksheet["C4"]
    period_cell.value = default_period
    period_cell.alignment = selector_alignment
    period_cell.fill = selector_fill

    year_validation = DataValidation(type="list", formula1="=YearsList", allow_blank=False)
    period_validation = DataValidation(type="list", formula1="=MonthsList", allow_blank=False)

    worksheet.add_data_validation(year_validation)
    worksheet.add_data_validation(period_validation)
    year_validation.add(year_cell.coordinate)
    period_validation.add(period_cell.coordinate)


def _build_kpi_tiles(
    worksheet: Worksheet,
    tracking_balance_formula: str,
    savings_rate_formula: str,
) -> None:
    label_font = Font(bold=True)
    value_alignment = Alignment(horizontal="center")
    value_fill = PatternFill(
        start_color=TILE_FILL,
        end_color=TILE_FILL,
        fill_type="solid",
    )
    border = Border(
        left=Side(style="thin", color="C5D1DE"),
        right=Side(style="thin", color="C5D1DE"),
        top=Side(style="thin", color="C5D1DE"),
        bottom=Side(style="thin", color="C5D1DE"),
    )

    labels = (
        ("B6", "Selected Year"),
        ("B7", "Selected Period"),
        ("B8", "Tracking Balance"),
        ("B9", "Savings Rate"),
    )
    for cell_ref, label in labels:
        worksheet[cell_ref].value = label
        worksheet[cell_ref].font = label_font

    worksheet["C6"].value = "=DashYear"
    worksheet["C7"].value = "=DashPeriod"
    worksheet["C8"].value = tracking_balance_formula
    worksheet["C9"].value = savings_rate_formula

    for coord in ("C6", "C7", "C8", "C9"):
        cell = worksheet[coord]
        cell.alignment = value_alignment
        cell.fill = value_fill

    worksheet["C8"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    worksheet["C9"].number_format = "0.0%"

    for row in range(6, 10):
        for col in range(2, 4):
            worksheet.cell(row=row, column=col).border = border


def register_dashboard_named_ranges(manager: NamedRangeManager) -> None:
    """Register dashboard-specific named ranges."""

    specs = (
        NamedRangeSpec("DashYear", "Budget Dashboard", "$C$3"),
        NamedRangeSpec("DashPeriod", "Budget Dashboard", "$C$4"),
    )
    manager.register_many(specs)
