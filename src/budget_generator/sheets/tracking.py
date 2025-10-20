"""Budget Tracking worksheet builder with validations and formulas."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from typing import Mapping, Sequence

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


HEADERS: tuple[str, ...] = (
    "Date",
    "Type",
    "Category",
    "Amount",
    "Details",
    "Balance",
    "Effective Date",
)

ACCOUNTING_FORMAT = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
DATE_FORMAT = "yyyy-mm-dd"


@dataclass(frozen=True)
class TrackingEntry:
    date: datetime
    transaction_type: str
    category: str
    amount: float
    details: str | None = None


@dataclass
class TrackingConfig:
    """Configuration for building the tracking sheet."""

    max_rows: int = 200
    table_name: str = "tblTracking"
    header_row: int = 11
    start_column: int = 3  # column C
    intro_title: str = "Budget Tracking"
    intro_duration: str = "1h 33min"
    tutorial_note: str = "Tutorial at 1h 14min"
    pause_note: str = "Parei at 1h 14min "
    sample_entries: tuple[TrackingEntry, ...] = ()

    @property
    def data_start_row(self) -> int:
        return self.header_row + 1

    @property
    def end_row(self) -> int:
        return self.max_rows

    @property
    def end_column(self) -> int:
        return self.start_column + len(HEADERS) - 1

    @property
    def table_ref(self) -> str:
        start_letter = get_column_letter(self.start_column)
        end_letter = get_column_letter(self.end_column)
        return f"{start_letter}{self.header_row}:{end_letter}{self.end_row}"


def build_tracking_sheet(worksheet: Worksheet, spec: Mapping[str, object] | None = None) -> None:
    """Build the Budget Tracking sheet end-to-end."""

    config = _resolve_config(spec)
    _apply_intro_content(worksheet, config)
    _render_headers(worksheet, config)
    _set_column_widths(worksheet)
    _populate_sample_entries(worksheet, config)
    _create_table(worksheet, config)
    _apply_number_formats(worksheet, config)
    add_tracking_validations(worksheet, config)
    add_tracking_formulas(worksheet, config)
    add_tracking_conditional_formatting(worksheet, config)


def add_tracking_validations(worksheet: Worksheet, config: TrackingConfig | None = None) -> None:
    """Attach date/type/category validations required by the PRD."""

    cfg = config or TrackingConfig()

    date_validation = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2000,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
    )
    worksheet.add_data_validation(date_validation)
    first_col_letter = get_column_letter(cfg.start_column)
    date_validation.add(
        f"{first_col_letter}{cfg.data_start_row}:{first_col_letter}{cfg.end_row}"
    )

    type_validation = DataValidation(
        type="list",
        formula1='"Income,Expense,Saving"',
        allow_blank=False,
    )
    worksheet.add_data_validation(type_validation)
    type_col_letter = get_column_letter(cfg.start_column + 1)
    type_validation.add(
        f"{type_col_letter}{cfg.data_start_row}:{type_col_letter}{cfg.end_row}"
    )

    for row in range(cfg.data_start_row, cfg.end_row + 1):
        category_letter = get_column_letter(cfg.start_column + 2)
        formula = (
            f'=IF(${type_col_letter}{row}="Income",IncomeCats,'
            f'IF(${type_col_letter}{row}="Expense",ExpenseCats,SavingsCats))'
        )
        category_validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        worksheet.add_data_validation(category_validation)
        category_validation.add(f"{category_letter}{row}")


def add_tracking_formulas(worksheet: Worksheet, config: TrackingConfig | None = None) -> None:
    """Populate balance and effective-date formulas."""

    cfg = config or TrackingConfig()
    for row in range(cfg.data_start_row, cfg.end_row + 1):
        balance_cell = worksheet.cell(row=row, column=cfg.start_column + 5)
        balance_cell.value = (
            "=SUMPRODUCT((tblTracking[Date]<=[@Date])*(tblTracking[Type]=\"Income\")*"
            "tblTracking[Amount])"
            "-SUMPRODUCT((tblTracking[Date]<=[@Date])*((tblTracking[Type]=\"Expense\")+"
            "(tblTracking[Type]=\"Saving\"))*tblTracking[Amount])"
        )
        balance_cell.number_format = ACCOUNTING_FORMAT

        effective_cell = worksheet.cell(row=row, column=cfg.start_column + 6)
        effective_cell.value = (
            "=IF(AND(LateIncomeEnabled,[@Type]=\"Income\",DAY([@Date])>LateIncomeDay),"
            "DATE(YEAR([@Date]),MONTH([@Date])+1,1),[@Date])"
        )
        effective_cell.number_format = DATE_FORMAT


def add_tracking_conditional_formatting(
    worksheet: Worksheet, config: TrackingConfig | None = None
) -> None:
    """Apply conditional formatting rules called out in the PRD."""

    cfg = config or TrackingConfig()
    start_row = cfg.data_start_row
    end_row = cfg.end_row

    from openpyxl.formatting.rule import FormulaRule

    category_letter = get_column_letter(cfg.start_column + 2)
    type_letter = get_column_letter(cfg.start_column + 1)
    cat_range = f"{category_letter}{start_row}:{category_letter}{end_row}"
    worksheet.conditional_formatting.add(
        cat_range,
        FormulaRule(
            formula=[f"ISNA({category_letter}{start_row})"],
            fill=PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid"),
        ),
    )

    amt_range = f"{category_letter}{start_row}:{category_letter}{end_row}"
    worksheet.conditional_formatting.add(
        amt_range,
        FormulaRule(
            formula=[f"${type_letter}{start_row}=\"Income\""],
            fill=PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        ),
    )


def _resolve_config(spec: Mapping[str, object] | None) -> TrackingConfig:
    spec = spec or {}
    max_rows = int(spec.get("max_rows", 200))

    intro = spec.get("intro", {}) if isinstance(spec, Mapping) else {}
    notes = spec.get("notes", {}) if isinstance(spec, Mapping) else {}
    entries_spec: Sequence[Mapping[str, object]] = ()
    if isinstance(spec, Mapping):
        entries_spec = spec.get("sample_entries", ())  # type: ignore[assignment]

    sample_entries = _coerce_entries(entries_spec)
    if not sample_entries:
        sample_entries = (
            TrackingEntry(
                date=datetime(2017, 1, 1),
                transaction_type="Income",
                category="DiDi",
                amount=7700.5,
                details=None,
            ),
            TrackingEntry(
                date=datetime(2017, 3, 2),
                transaction_type="Savings",
                category="ETFs",
                amount=5000,
                details=None,
            ),
            TrackingEntry(
                date=datetime(2017, 3, 3),
                transaction_type="Expenses",
                category="Groceries",
                amount=500,
                details=None,
            ),
        )

    return TrackingConfig(
        max_rows=max_rows,
        intro_title=str(intro.get("title", "Budget Tracking")),
        intro_duration=str(intro.get("duration", "1h 33min")),
        tutorial_note=str(notes.get("tutorial_label", "Tutorial at 1h 14min")),
        pause_note=str(notes.get("pause_label", "Parei at 1h 14min ")),
        sample_entries=sample_entries,
    )


def _render_headers(worksheet: Worksheet, config: TrackingConfig) -> None:
    header_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")

    for offset, header in enumerate(HEADERS):
        column = config.start_column + offset
        cell = worksheet.cell(row=config.header_row, column=column, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill


def _set_column_widths(worksheet: Worksheet) -> None:
    widths = {
        "B": 40,
        "C": 14,
        "D": 12,
        "E": 24,
        "F": 12,
        "G": 30,
        "H": 16,
        "I": 16,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _create_table(worksheet: Worksheet, config: TrackingConfig) -> None:
    table = Table(displayName=config.table_name, ref=config.table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _apply_intro_content(worksheet: Worksheet, config: TrackingConfig) -> None:
    """Write descriptive header content above the tracking table."""

    title_cell = worksheet["B1"]
    title_cell.value = config.intro_title
    title_cell.font = Font(bold=True, size=16)

    duration_cell = worksheet["E5"]
    duration_cell.value = config.intro_duration
    duration_cell.font = Font(italic=True)

    worksheet["B6"].value = config.tutorial_note
    worksheet["B7"].value = config.pause_note


def _populate_sample_entries(worksheet: Worksheet, config: TrackingConfig) -> None:
    """Insert illustrative rows that match the expected design."""

    for offset, entry in enumerate(config.sample_entries):
        row = config.data_start_row + offset
        if row > config.end_row:
            break

        worksheet.cell(row=row, column=config.start_column, value=entry.date)
        worksheet.cell(
            row=row,
            column=config.start_column + 1,
            value=entry.transaction_type,
        )
        worksheet.cell(row=row, column=config.start_column + 2, value=entry.category)
        worksheet.cell(row=row, column=config.start_column + 3, value=entry.amount)
        if entry.details:
            worksheet.cell(row=row, column=config.start_column + 4, value=entry.details)


def _apply_number_formats(worksheet: Worksheet, config: TrackingConfig) -> None:
    for row in range(config.data_start_row, config.end_row + 1):
        date_cell = worksheet.cell(row=row, column=config.start_column)
        date_cell.number_format = DATE_FORMAT

        amount_cell = worksheet.cell(row=row, column=config.start_column + 3)
        amount_cell.number_format = ACCOUNTING_FORMAT

        details_cell = worksheet.cell(row=row, column=config.start_column + 4)
        details_cell.number_format = "@"


def _coerce_entries(
    entries: Sequence[Mapping[str, object]]
) -> tuple[TrackingEntry, ...]:
    """Convert raw mapping data into :class:`TrackingEntry` records."""

    coerced: list[TrackingEntry] = []
    for entry in entries:
        if not isinstance(entry, Mapping):
            continue

        when = _coerce_datetime(entry.get("date"))
        transaction_type = entry.get("type")
        category = entry.get("category")
        amount = entry.get("amount")

        if when is None or not transaction_type or not category or amount is None:
            continue

        details_value = entry.get("details")
        coerced.append(
            TrackingEntry(
                date=when,
                transaction_type=str(transaction_type),
                category=str(category),
                amount=float(amount),
                details=str(details_value) if details_value not in (None, "") else None,
            )
        )

    return tuple(coerced)


def _coerce_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return None
    return None
