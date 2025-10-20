"""Budget Planning worksheet builder and helpers."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable, Mapping

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..formatting.conditional import add_unallocated_conditional_formatting
from ..utils.named_ranges import NamedRangeManager, NamedRangeSpec


MONTHS: tuple[str, ...] = (
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
)


ACCOUNTING_FORMAT = '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'


@dataclass(frozen=True)
class SectionDefinition:
    """Metadata that describes a planning sheet section."""

    key: str
    title: str
    title_row: int
    start_row: int
    total_row: int
    fill_color: str
    categories: Iterable[str]


@dataclass(frozen=True)
class YearBlock:
    """Description of a single year scaffold block in the planning sheet."""

    offset: int
    month_columns: tuple[int, ...]
    total_column: int

    @property
    def start_column(self) -> int:
        return self.month_columns[0]


class PlanningSheetBuilder:
    """Internal helper that encapsulates worksheet layout and styling."""

    CATEGORY_COLUMN = 4  # Column D
    YEAR_START_COLUMN = 5  # Column E
    GAP_BETWEEN_BLOCKS = 1
    UNALLOCATED_ROW = 7

    SECTION_DEFINITIONS: tuple[SectionDefinition, ...] = (
        SectionDefinition(
            key="income",
            title="Income",
            title_row=10,
            start_row=12,
            total_row=24,
            fill_color="43D40F",
            categories=(
                "Salary",
                "Freelance",
                "Investments",
                "Other",
            ),
        ),
        SectionDefinition(
            key="expenses",
            title="Expenses",
            title_row=31,
            start_row=33,
            total_row=45,
            fill_color="F01010",
            categories=(
                "Housing",
                "Utilities",
                "Groceries",
                "Transportation",
                "Insurance",
                "Healthcare",
                "Debt Repayments",
                "Entertainment",
                "Subscriptions",
                "Miscellaneous",
            ),
        ),
        SectionDefinition(
            key="savings",
            title="Savings",
            title_row=53,
            start_row=55,
            total_row=67,
            fill_color="1564ED",
            categories=(
                "Emergency Fund",
                "Retirement",
                "Investments",
                "Vacation",
                "Other",
            ),
        ),
    )

    def __init__(self, worksheet: Worksheet, spec: Mapping[str, object]):
        self.ws = worksheet
        self.spec = spec
        self.scaffold_years = max(1, int(self.spec.get("scaffold_years", 16)))
        self.year_blocks: list[YearBlock] = []
        self.year_block_width = len(MONTHS) + 1 + self.GAP_BETWEEN_BLOCKS

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def build(self) -> None:
        self._build_hero_header()
        self._build_year_blocks()
        self._label_unallocated_row()
        for section in self.SECTION_DEFINITIONS:
            self._render_section(section)
        self._populate_unallocated_formulas()
        self._apply_conditional_formatting()
        freeze_column = get_column_letter(self.year_blocks[0].start_column)
        self.ws.freeze_panes = f"{freeze_column}12"

    # ------------------------------------------------------------------
    # Hero + year scaffolding
    # ------------------------------------------------------------------
    def _build_hero_header(self) -> None:
        title = self.spec.get("hero_title", "Budget Planning")
        subtitle = self.spec.get(
            "hero_subtitle",
            "Plan your income, spending, and savings across the year.",
        )

        hero_cell = self.ws["C1"]
        hero_cell.value = title
        hero_cell.font = Font(bold=True, size=16)

        subtitle_cell = self.ws["C3"]
        subtitle_cell.value = subtitle
        subtitle_cell.font = Font(italic=True, size=11)
        subtitle_cell.alignment = Alignment(wrap_text=True)

    def _build_year_blocks(self) -> None:
        self.year_blocks.clear()
        for offset in range(self.scaffold_years):
            block = self._build_year_block(offset)
            self.year_blocks.append(block)

    def _build_year_block(self, offset: int) -> YearBlock:
        start_col = self.YEAR_START_COLUMN + offset * self.year_block_width
        month_columns = tuple(range(start_col, start_col + len(MONTHS)))
        total_column = start_col + len(MONTHS)

        # Year banner (row 5)
        banner_cell = self.ws.cell(row=5, column=start_col)
        if offset == 0:
            banner_cell.value = "=starting_year"
        else:
            previous_column = start_col - self.year_block_width
            previous_letter = get_column_letter(previous_column)
            banner_cell.value = f"={previous_letter}5+1"
        banner_cell.font = Font(bold=True, size=13)
        banner_cell.alignment = Alignment(horizontal="center")
        banner_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
        banner_cell.fill = banner_fill
        if total_column > start_col:
            self.ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=total_column)

        header_fill = PatternFill(start_color="DAE3F3", end_color="DAE3F3", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")

        for column, month in zip(month_columns, MONTHS):
            letter = get_column_letter(column)
            header_cell = self.ws.cell(row=6, column=column)
            header_cell.value = f'=IF({letter}{self.UNALLOCATED_ROW}=0,"{month} ✓","{month}")'
            header_cell.font = header_font
            header_cell.alignment = header_alignment
            header_cell.fill = header_fill

        total_letter = get_column_letter(total_column)
        total_header = self.ws.cell(row=6, column=total_column)
        total_header.value = f'=IF({total_letter}{self.UNALLOCATED_ROW}=0,"Total ✓","Total")'
        total_header.font = header_font
        total_header.alignment = header_alignment
        total_header.fill = header_fill

        note_row = 8
        note_cell = self.ws.cell(row=note_row, column=start_col)
        note_cell.value = (
            "Year 1 overview" if offset == 0 else f"Year {offset + 1} scaffold – extend rows as needed"
        )
        note_cell.font = Font(size=10, italic=True)
        note_cell.alignment = Alignment(wrap_text=True)

        return YearBlock(offset=offset, month_columns=month_columns, total_column=total_column)

    # ------------------------------------------------------------------
    # Sections
    # ------------------------------------------------------------------
    def _render_section(self, section: SectionDefinition) -> None:
        title_cell = self.ws.cell(row=section.title_row, column=self.CATEGORY_COLUMN)
        title_cell.value = section.title
        title_cell.font = Font(bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(
            start_color=section.fill_color, end_color=section.fill_color, fill_type="solid"
        )
        title_cell.alignment = Alignment(horizontal="left")

        self._initialise_category_rows(section)
        self._write_section_totals(section)
        self._apply_section_borders(section)

    def _initialise_category_rows(self, section: SectionDefinition) -> None:
        row_count = section.total_row - section.start_row
        categories = list(section.categories)
        if len(categories) < row_count:
            categories.extend([""] * (row_count - len(categories)))

        for offset, category in enumerate(categories):
            row = section.start_row + offset
            category_cell = self.ws.cell(row=row, column=self.CATEGORY_COLUMN)
            category_cell.value = category

            for block in self.year_blocks:
                for column in block.month_columns:
                    cell = self.ws.cell(row=row, column=column)
                    cell.value = 0
                    cell.number_format = ACCOUNTING_FORMAT

                total_cell = self.ws.cell(row=row, column=block.total_column)
                total_cell.value = self._row_total_formula(row, block)
                total_cell.number_format = ACCOUNTING_FORMAT

    def _write_section_totals(self, section: SectionDefinition) -> None:
        total_label = self.ws.cell(row=section.total_row, column=self.CATEGORY_COLUMN)
        total_label.value = f"Total {section.title}"
        total_label.font = Font(bold=True)
        total_label.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True)

        for block in self.year_blocks:
            for column in block.month_columns:
                cell = self.ws.cell(row=section.total_row, column=column)
                cell.value = self._section_total_formula(section)
                cell.number_format = ACCOUNTING_FORMAT
                cell.font = total_font
                cell.fill = total_fill

            total_cell = self.ws.cell(row=section.total_row, column=block.total_column)
            total_cell.value = self._section_total_formula(section)
            total_cell.number_format = ACCOUNTING_FORMAT
            total_cell.font = total_font
            total_cell.fill = total_fill

    def _apply_section_borders(self, section: SectionDefinition) -> None:
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        end_column = self.year_blocks[-1].total_column
        for row in range(section.title_row, section.total_row + 1):
            for column in range(self.CATEGORY_COLUMN, end_column + 1):
                self.ws.cell(row=row, column=column).border = border

    # ------------------------------------------------------------------
    # Unallocated row
    # ------------------------------------------------------------------
    def _label_unallocated_row(self) -> None:
        label_cell = self.ws.cell(row=self.UNALLOCATED_ROW, column=self.CATEGORY_COLUMN)
        label_cell.value = "Unallocated (per month)"
        label_cell.font = Font(bold=True)

    def _populate_unallocated_formulas(self) -> None:
        income_total_row = self.SECTION_DEFINITIONS[0].total_row
        expense_total_row = self.SECTION_DEFINITIONS[1].total_row
        savings_total_row = self.SECTION_DEFINITIONS[2].total_row

        for block in self.year_blocks:
            for column in (*block.month_columns, block.total_column):
                letter = get_column_letter(column)
                income_ref = f"{letter}{income_total_row}"
                expense_ref = f"{letter}{expense_total_row}"
                savings_ref = f"{letter}{savings_total_row}"
                cell = self.ws.cell(row=self.UNALLOCATED_ROW, column=column)
                cell.value = f"={income_ref}-({expense_ref}+{savings_ref})"
                cell.number_format = ACCOUNTING_FORMAT
                cell.font = Font(bold=True)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _row_total_formula(self, row: int, block: YearBlock) -> str:
        start_letter = get_column_letter(block.month_columns[0])
        end_letter = get_column_letter(block.month_columns[-1])
        return f"=SUM({start_letter}{row}:{end_letter}{row})"

    def _section_total_formula(self, section: SectionDefinition) -> str:
        min_name = f"{section.key}_min_row"
        max_name = f"{section.key}_max_row"
        return (
            "=SUM(INDIRECT(ADDRESS("
            f"{min_name},COLUMN()) & \":\" & ADDRESS({max_name},COLUMN())))"
        )

    def _apply_conditional_formatting(self) -> None:
        start_letter = get_column_letter(self.year_blocks[0].start_column)
        end_letter = get_column_letter(self.year_blocks[-1].total_column)
        add_unallocated_conditional_formatting(
            self.ws, start_letter, end_letter, self.UNALLOCATED_ROW
        )


def build_planning_sheet(worksheet: Worksheet, spec: Mapping[str, object] | None = None) -> None:
    builder = PlanningSheetBuilder(worksheet, spec or {})
    builder.build()


def register_planning_named_ranges(manager: NamedRangeManager) -> None:
    """Register named ranges for the Budget-Planning sheet."""

    specs = [
        NamedRangeSpec("IncomeCats", "Budget-Planning", "$D$12:$D$23"),
        NamedRangeSpec("ExpenseCats", "Budget-Planning", "$D$33:$D$44"),
        NamedRangeSpec("SavingsCats", "Budget-Planning", "$D$55:$D$66"),
        NamedRangeSpec("IncomeGrid", "Budget-Planning", "$E$12:$Q$23"),
        NamedRangeSpec("ExpenseGrid", "Budget-Planning", "$E$33:$Q$44"),
        NamedRangeSpec("SavingsGrid", "Budget-Planning", "$E$55:$Q$66"),
        NamedRangeSpec("IncomeHeader", "Budget-Planning", "$D$10"),
        NamedRangeSpec("ExpenseHeader", "Budget-Planning", "$D$31"),
        NamedRangeSpec("SavingsHeader", "Budget-Planning", "$D$53"),
        NamedRangeSpec("IncomeTotals", "Budget-Planning", "$E$24:$Q$24"),
        NamedRangeSpec("ExpenseTotals", "Budget-Planning", "$E$45:$Q$45"),
        NamedRangeSpec("SavingsTotals", "Budget-Planning", "$E$67:$Q$67"),
        NamedRangeSpec("UnallocatedRow", "Budget-Planning", "$E$7:$Q$7"),
    ]

    manager.register_many(specs)

    constants = {
        "income_min_row": 12,
        "income_max_row": 23,
        "expenses_min_row": 33,
        "expenses_max_row": 44,
        "savings_min_row": 55,
        "savings_max_row": 66,
    }

    for name, value in constants.items():
        manager.create_constant(name, value)
