"""Conditional formatting helpers."""

from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def add_unallocated_conditional_formatting(
    worksheet: Worksheet, start_col: str, end_col: str, row: int
) -> None:
    """Attach the three-state colouring rules for the Unallocated row."""

    range_str = f"{start_col}{row}:{end_col}{row}"

    green_fill = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")
    rule_equal_zero = CellIsRule(operator="equal", formula=["0"], fill=green_fill)

    red_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    rule_less_than = CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)

    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    formula = f"AND({start_col}13=0,{start_col}26=0,{start_col}34=0)"
    rule_all_zero = FormulaRule(formula=[formula], fill=gray_fill)

    worksheet.conditional_formatting.add(range_str, rule_equal_zero)
    worksheet.conditional_formatting.add(range_str, rule_less_than)
    worksheet.conditional_formatting.add(range_str, rule_all_zero)
