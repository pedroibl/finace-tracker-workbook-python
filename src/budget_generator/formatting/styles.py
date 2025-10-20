"""Reusable styling helpers for worksheet builders."""

from __future__ import annotations

from typing import Optional

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def apply_fill(cell, hex_color: str):
    """Apply a solid fill to *cell* using a hex colour (#RRGGBB or RRGGBB)."""

    colour_value = hex_color.lstrip("#").upper()
    cell.fill = PatternFill(
        start_color=colour_value,
        end_color=colour_value,
        fill_type="solid",
    )
    return cell


def merge_and_format(
    worksheet: Worksheet,
    cell_range: str,
    *,
    value: str,
    font: Optional[Font] = None,
    alignment: Optional[Alignment] = None,
    fill_color: Optional[str] = None,
):
    """Merge *cell_range* and apply value plus optional formatting in one step."""

    worksheet.merge_cells(cell_range)
    anchor = worksheet[cell_range.split(":")[0]]  # top-left cell after merge
    anchor.value = value
    if font is not None:
        anchor.font = font
    if alignment is not None:
        anchor.alignment = alignment
    if fill_color is not None:
        apply_fill(anchor, fill_color)
    return anchor
