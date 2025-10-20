"""Output validation comparisons against a golden workbook."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from budget_generator.generator import BudgetGenerator
from budget_generator.utils.json_loader import load_json_spec

GOLDEN_PATH = Path("tests/fixtures/golden_tutorial.xlsx")
SPEC_PATH = Path("examples/tutorial_spec.json")


@pytest.mark.output
def test_generated_workbook_matches_golden(tmp_path: Path) -> None:
    if not GOLDEN_PATH.exists():
        pytest.skip("Golden workbook not yet created")

    output_path = tmp_path / "tutorial_generated.xlsx"
    spec = load_json_spec(SPEC_PATH)
    generator = BudgetGenerator(spec)
    generator.create_workbook()
    generator.create_sheets()
    generator.build_sheet_contents()
    generator.save_workbook(output_path)

    golden = openpyxl.load_workbook(GOLDEN_PATH)
    generated = openpyxl.load_workbook(output_path)
    try:
        assert golden.sheetnames == generated.sheetnames

        for sheet_name in golden.sheetnames:
            golden_ws = golden[sheet_name]
            generated_ws = generated[sheet_name]

            golden_cells = list(golden_ws.iter_rows(values_only=True))
            generated_cells = list(generated_ws.iter_rows(values_only=True))
            assert golden_cells == generated_cells, f"Mismatch in {sheet_name} values"

            if sheet_name == "Budget Dashboard":
                assert len(golden_ws._charts) == len(generated_ws._charts)
            if sheet_name == "Budget Planning":
                assert golden_ws["D13"].value == generated_ws["D13"].value
                assert golden_ws["D13"].number_format == generated_ws["D13"].number_format
            if sheet_name == "Budget Tracking":
                golden_table = next(iter(golden_ws.tables.values()))
                generated_table = next(iter(generated_ws.tables.values()))
                assert golden_table.ref == generated_table.ref
                assert golden_ws["B3"].number_format == generated_ws["B3"].number_format
    finally:
        golden.close()
        generated.close()
