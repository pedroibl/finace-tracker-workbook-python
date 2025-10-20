from __future__ import annotations

from openpyxl import Workbook

from budget_generator.sheets.tracking import (
    TrackingConfig,
    add_tracking_conditional_formatting,
    add_tracking_formulas,
    add_tracking_validations,
    build_tracking_sheet,
)


def build_sheet(max_rows: int = 50):
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Tracking"
    build_tracking_sheet(ws, {"max_rows": max_rows})
    return ws


def test_tracking_table_structure() -> None:
    ws = build_sheet(20)

    assert "tblTracking" in ws.tables
    table = ws.tables["tblTracking"]
    assert table.ref == "B2:H20"

    headers = [ws.cell(row=2, column=col).value for col in range(2, 9)]
    assert headers == [
        "Date",
        "Type",
        "Category",
        "Amount",
        "Details",
        "Balance",
        "Effective Date",
    ]

    assert ws.column_dimensions["B"].width == 14
    assert ws["B3"].number_format == "yyyy-mm-dd"
    assert ws["D3"].number_format.startswith("_($*")


def test_tracking_validations_created() -> None:
    cfg = TrackingConfig(max_rows=12)
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Tracking"
    build_tracking_sheet(ws, {"max_rows": 12})

    validations = list(ws.data_validations.dataValidation)
    assert any(v.type == "date" for v in validations)
    assert any(v.type == "list" and v.formula1 == '"Income,Expense,Saving"' for v in validations)

    category_validations = [v for v in validations if "IncomeCats" in v.formula1]
    assert len(category_validations) == cfg.max_rows - (cfg.header_row)


def test_tracking_formulas_and_conditional_formatting() -> None:
    ws = build_sheet(8)

    assert (
        ws["G3"].value
        == "=SUMPRODUCT((tblTracking[Date]<=[@Date])*(tblTracking[Type]=\"Income\")*tblTracking[Amount])-SUMPRODUCT((tblTracking[Date]<=[@Date])*((tblTracking[Type]=\"Expense\")+(tblTracking[Type]=\"Saving\"))*tblTracking[Amount])"
    )
    assert (
        ws["H3"].value
        == "=IF(AND(LateIncomeEnabled,[@Type]=\"Income\",DAY([@Date])>LateIncomeDay),DATE(YEAR([@Date]),MONTH([@Date])+1,1),[@Date])"
    )

    rules = []
    for cf in ws.conditional_formatting:
        rules.extend(cf.rules)

    assert any(rule.type == "expression" and "ISNA" in rule.formula[0] for rule in rules)
    assert any(rule.type == "expression" and "Income" in rule.formula[0] for rule in rules)
