"""Tests for the Calculations worksheet builder."""

from __future__ import annotations

from openpyxl import Workbook

from budget_generator.sheets.calculations import build_calculations_sheet
from budget_generator.sheets.planning import ACCOUNTING_FORMAT


def _build_sheet() -> tuple[Workbook, str]:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Calculations"
    build_calculations_sheet(worksheet, {})
    return workbook, worksheet.title


def test_metric_tiles_headers_and_formulas() -> None:
    workbook, sheet_name = _build_sheet()
    ws = workbook[sheet_name]

    headers = [ws.cell(row=2, column=col).value for col in range(2, 5)]
    assert headers == ["Metric", "Value", "Notes"]
    assert ws["B2"].fill.start_color.rgb[-6:] == "EAD1DC"

    assert ws["B3"].value == "Current Date"
    assert ws["C3"].value == "=TODAY()"
    assert ws["C3"].number_format == "yyyy-mm-dd"

    assert ws["B4"].value == "Last Record Date"
    assert ws["C4"].value == "=MAX(tblTracking[Date])"

    assert ws["B5"].value == "Number of Records"
    assert ws["C5"].value == "=COUNTA(tblTracking[Date])"

    assert (
        ws["C6"].value
        == "=IFERROR(LOOKUP(2,1/(tblTracking[Date]<>\"\"),tblTracking[Balance]),0)"
    )
    assert ws["C6"].number_format == ACCOUNTING_FORMAT


def test_month_map_and_idx_formula() -> None:
    workbook, sheet_name = _build_sheet()
    ws = workbook[sheet_name]

    months = [ws.cell(row=row, column=10).value for row in range(2, 14)]
    assert months == [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]

    indices = [ws.cell(row=row, column=11).value for row in range(2, 14)]
    assert indices == list(range(1, 13))

    assert (
        ws["K1"].value
        == "=INDEX(INDEX(MonthMap,0,2),MATCH(DashPeriod,INDEX(MonthMap,0,1),0))"
    )


def test_budget_vs_tracked_table_formulas() -> None:
    workbook, sheet_name = _build_sheet()
    ws = workbook[sheet_name]

    headers = [ws.cell(row=2, column=col).value for col in range(5, 9)]
    assert headers == ["Section", "BudgetedMonth", "TrackedMonth", "Remaining"]
    assert ws["E2"].fill.start_color.rgb[-6:] == "DEEAF6"

    assert ws["F3"].value == "=CHOOSE(MonthIdx,D13,E13,F13,G13,H13,I13,J13,K13,L13,M13,N13,O13)"
    assert ws["G3"].value == (
        "=SUMPRODUCT((MONTH(tblTracking[Effective Date])=MonthIdx)*(tblTracking[Type]=\"Income\")*tblTracking[Amount])"
    )
    assert ws["H3"].value == "=F3-G3"

    assert ws["F4"].value == "=CHOOSE(MonthIdx,D26,E26,F26,G26,H26,I26,J26,K26,L26,M26,N26,O26)"
    assert ws["G4"].value == (
        "=SUMPRODUCT((MONTH(tblTracking[Effective Date])=MonthIdx)*(tblTracking[Type]=\"Expense\")*tblTracking[Amount])"
    )
    assert ws["H4"].value == "=F4-G4"

    assert ws["F5"].value == "=CHOOSE(MonthIdx,D34,E34,F34,G34,H34,I34,J34,K34,L34,M34,N34,O34)"
    assert ws["G5"].value == (
        "=SUMPRODUCT((MONTH(tblTracking[Effective Date])=MonthIdx)*(tblTracking[Type]=\"Saving\")*tblTracking[Amount])"
    )
    assert ws["H5"].value == "=F5-G5"

    for column in (6, 7, 8):
        assert ws.cell(row=3, column=column).number_format == ACCOUNTING_FORMAT
