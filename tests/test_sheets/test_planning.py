from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from budget_generator.sheets.planning import ACCOUNTING_FORMAT, MONTHS, build_planning_sheet


def _year_block_start(block_index: int) -> int:
    """Return the starting column index for the given 1-based *block_index*."""

    block_width = len(MONTHS) + 2  # 12 months + total + gap
    return 5 + (block_index - 1) * block_width


def test_planning_banner_and_headers() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget-Planning"

    build_planning_sheet(ws, {})

    assert ws["C1"].value == "Budget Planning"
    assert ws["C3"].value.startswith("Plan your")
    assert ws["E5"].value == "=starting_year"
    assert ws["E6"].value == '=IF(E7=0,"Jan ✓","Jan")'
    assert ws["Q6"].value == '=IF(Q7=0,"Total ✓","Total")'


def test_planning_sections_and_totals() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget-Planning"

    build_planning_sheet(ws, {})

    assert ws["D10"].value == "Income"
    assert ws["D10"].fill.start_color.rgb[-6:] == "43D40F"
    assert ws["D12"].value == "Salary"
    assert ws["D24"].value == "Total Income"
    expected_income_total = (
        '=SUM(INDIRECT(ADDRESS(income_min_row,COLUMN()) & ":" & ADDRESS(income_max_row,COLUMN())))'
    )
    assert ws["E24"].value == expected_income_total

    assert ws["D31"].value == "Expenses"
    assert ws["D31"].fill.start_color.rgb[-6:] == "F01010"
    assert ws["D33"].value == "Housing"
    assert ws["D45"].value == "Total Expenses"

    assert ws["D53"].value == "Savings"
    assert ws["D53"].fill.start_color.rgb[-6:] == "1564ED"
    assert ws["D55"].value == "Emergency Fund"
    assert ws["D67"].value == "Total Savings"

    for column in range(5, 17):  # Columns E through P
        assert ws.cell(row=12, column=column).value == 0
        assert ws.cell(row=12, column=column).number_format == ACCOUNTING_FORMAT

    total_cell = ws.cell(row=12, column=17)
    assert total_cell.value == "=SUM(E12:P12)"
    assert total_cell.number_format == ACCOUNTING_FORMAT

    # Subsequent year blocks should also initialise rows and totals
    second_block_start = _year_block_start(2)
    second_block_end = second_block_start + len(MONTHS) - 1
    for column in range(second_block_start, second_block_end + 1):
        assert ws.cell(row=12, column=column).value == 0
        assert ws.cell(row=12, column=column).number_format == ACCOUNTING_FORMAT

    second_total_column = second_block_start + len(MONTHS)
    start_letter = get_column_letter(second_block_start)
    end_letter = get_column_letter(second_block_end)
    expected_formula = f"=SUM({start_letter}12:{end_letter}12)"
    assert ws.cell(row=12, column=second_total_column).value == expected_formula
    assert (
        ws.cell(row=12, column=second_total_column).number_format == ACCOUNTING_FORMAT
    )


def test_unallocated_row_formulas_and_conditional_formatting() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget-Planning"

    build_planning_sheet(ws, {})

    for column in range(5, 18):
        col_letter = ws.cell(row=7, column=column).column_letter
        assert ws.cell(row=7, column=column).value == (
            f"={col_letter}24-({col_letter}45+{col_letter}67)"
        )

    # Later year blocks should mirror the same formula pattern
    second_block_start = _year_block_start(2)
    second_letter = get_column_letter(second_block_start)
    assert ws.cell(row=7, column=second_block_start).value == (
        f"={second_letter}24-({second_letter}45+{second_letter}67)"
    )

    final_total_column = _year_block_start(16) + len(MONTHS)
    final_letter = get_column_letter(final_total_column)
    assert ws.cell(row=7, column=final_total_column).value == (
        f"={final_letter}24-({final_letter}45+{final_letter}67)"
    )

    cf_rules = []
    for cf in ws.conditional_formatting:
        cf_rules.extend(cf.rules)

    assert any(rule.type == "cellIs" and rule.operator == "equal" for rule in cf_rules)
    assert any(rule.type == "cellIs" and rule.operator == "lessThan" for rule in cf_rules)
    assert any(rule.type == "expression" for rule in cf_rules)


def test_year_two_scaffold_present() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget-Planning"

    build_planning_sheet(ws, {})

    second_block_start = _year_block_start(2)
    second_banner = ws.cell(row=5, column=second_block_start)
    assert second_banner.value == "=E5+1"

    header_columns = range(second_block_start, second_block_start + len(MONTHS) + 1)
    second_year_headers = [ws.cell(row=6, column=col).value for col in header_columns]
    assert second_year_headers[0] == '=IF(S7=0,"Jan ✓","Jan")'
    assert ws.cell(row=8, column=second_block_start).value == (
        "Year 2 scaffold – extend rows as needed"
    )


def test_scaffold_years_config_creates_additional_headers() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget-Planning"

    build_planning_sheet(ws, {"scaffold_years": 3})

    # Year 3 banner should chain from the previous year and begin at column AG (index 33)
    assert ws["AG5"].value == "=S5+1"
    headers_year3 = [ws.cell(row=6, column=col).value for col in range(33, 46)]
    assert headers_year3[0] == '=IF(AG7=0,"Jan ✓","Jan")'
