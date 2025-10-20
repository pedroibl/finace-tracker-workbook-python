"""Tests for the Budget Dashboard worksheet builder."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.chart import DoughnutChart

from budget_generator.charts import add_dashboard_doughnut_charts
from budget_generator.sheets.calculations import build_calculations_sheet
from budget_generator.sheets.dashboard import build_dashboard_sheet


def _build_sheet(spec: dict | None = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Dashboard"
    build_dashboard_sheet(ws, spec or {})
    return wb, ws


def test_dashboard_headers_and_selectors() -> None:
    wb, ws = _build_sheet()

    headers = [ws.cell(row=2, column=col).value for col in range(2, 9)]
    assert headers == [
        "Year",
        "Period",
        "Category",
        "Tracked",
        "Budgeted",
        "% of Budget",
        "Remaining",
    ]
    assert ws["B2"].fill.start_color.rgb[-6:] == "DAEEF3"

    assert ws["B3"].value == "Year"
    assert ws["C3"].value == "=StartingYear"
    assert ws["C3"].fill.start_color.rgb[-6:] == "E7F3F9"
    assert ws["B4"].value == "Period"
    assert ws["C4"].value == "Jan"

    validations = list(ws.data_validations.dataValidation)
    assert any(v.formula1 == "=YearsList" for v in validations)
    assert any(v.formula1 == "=MonthsList" for v in validations)

    # KPI tiles
    assert ws["B6"].value == "Selected Year"
    assert ws["C6"].value == "=DashYear"
    assert ws["B8"].value == "Tracking Balance"
    assert ws["C8"].value == "=Calculations!C6"
    assert ws["C8"].number_format.startswith("_($*")
    assert ws["C9"].number_format == "0.0%"

    for coord in ("C6", "C7", "C8", "C9"):
        assert ws[coord].fill.start_color.rgb[-6:] == "F9FBFD"
        assert ws[coord].alignment.horizontal == "center"

    assert ws["C8"].border.left.style == "thin"
    assert ws["B9"].border.bottom.style == "thin"

    wb.close()


def test_dashboard_selector_defaults_can_be_customised() -> None:
    _, ws = _build_sheet(
        {
            "selectors": {
                "default_year_formula": "=StartingYear+1",
                "default_period": "Feb",
            }
        }
    )

    assert ws["C3"].value == "=StartingYear+1"
    assert ws["C4"].value == "Feb"


def test_dashboard_charts_created() -> None:
    wb = Workbook()
    calc_ws = wb.active
    calc_ws.title = "Calculations"
    build_calculations_sheet(calc_ws, {})

    dashboard_ws = wb.create_sheet("Budget Dashboard")
    build_dashboard_sheet(dashboard_ws, {})

    add_dashboard_doughnut_charts(dashboard_ws)

    charts = dashboard_ws._charts  # type: ignore[attr-defined]
    assert len(charts) == 3
    def title_string(chart: DoughnutChart) -> str:
        tx = chart.title
        if hasattr(tx, "tx") and tx.tx is not None:
            rich = tx.tx.rich
            if rich is not None and rich.p and rich.p[0].r:
                return "".join(r.t for r in rich.p[0].r if r.t)
        return str(chart.title)

    titles = [title_string(chart) for chart in charts]
    assert "Income (Budget vs Tracked)" in titles

    first_chart = charts[0]
    assert isinstance(first_chart, DoughnutChart)
    assert first_chart.series[0].dLbls.showPercent
    assert first_chart.series[0].dLbls.showVal

    wb.close()
