"""Integration tests for the end-to-end workbook generation."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from budget_generator.generator import BudgetGenerator, GeneratorError
from budget_generator.utils.json_loader import load_json_spec

EXAMPLES_DIR = Path("examples")
FIXTURES_DIR = Path("tests/fixtures")


@pytest.mark.integration
def test_full_workbook_generation(tmp_path: Path) -> None:
    spec_path = EXAMPLES_DIR / "tutorial_spec.json"
    output_path = tmp_path / "tutorial_output.xlsx"

    spec = load_json_spec(spec_path)
    generator = BudgetGenerator(spec)
    generator.create_workbook()
    generator.create_sheets()
    generator.build_sheet_contents()
    generator.save_workbook(output_path)

    assert output_path.exists()

    wb = openpyxl.load_workbook(output_path)
    try:
        assert wb.sheetnames == [
            "Settings",
            "Dropdown Data",
            "Budget Planning",
            "Budget Tracking",
            "Calculations",
            "Budget Dashboard",
        ]

        # Hidden helper sheets
        assert wb["Dropdown Data"].sheet_state == "hidden"
        assert wb["Calculations"].sheet_state == "hidden"

        defined_names = wb.defined_names
        expected_ranges = [
            "StartingYear",
            "LateIncomeEnabled",
            "YearsList",
            "MonthsList",
            "IncomeCats",
            "DashYear",
            "MonthIdx",
        ]
        for name in expected_ranges:
            assert name in defined_names, f"Named range {name} missing"

        planning_ws = wb["Budget Planning"]
        assert planning_ws["D13"].value.startswith("=SUM")
        tracking_ws = wb["Budget Tracking"]
        assert "tblTracking" in tracking_ws.tables

        dashboard_ws = wb["Budget Dashboard"]
        assert len(dashboard_ws._charts) == 3  # type: ignore[attr-defined]
    finally:
        wb.close()


@pytest.mark.integration
def test_generation_handles_empty_categories(tmp_path: Path) -> None:
    spec_path = FIXTURES_DIR / "empty_categories_spec.json"
    output_path = tmp_path / "empty_categories.xlsx"

    spec = load_json_spec(spec_path)
    generator = BudgetGenerator(spec)
    generator.create_workbook()
    generator.create_sheets()
    generator.build_sheet_contents()
    generator.save_workbook(output_path)

    wb = openpyxl.load_workbook(output_path)
    try:
        planning_ws = wb["Budget Planning"]
        # Income grid initialised to zeros even without explicit categories
        assert planning_ws["D8"].value == 0
        tracking_ws = wb["Budget Tracking"]
        assert tracking_ws["B3"].number_format == "yyyy-mm-dd"
    finally:
        wb.close()


@pytest.mark.integration
def test_create_sheets_fails_for_malformed_workbook() -> None:
    spec_path = FIXTURES_DIR / "malformed_workbook_spec.json"
    spec = load_json_spec(spec_path)
    generator = BudgetGenerator(spec)
    generator.create_workbook()

    with pytest.raises(GeneratorError):
        generator.create_sheets()
