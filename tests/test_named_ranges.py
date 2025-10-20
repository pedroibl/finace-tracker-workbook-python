from __future__ import annotations

import pytest
from openpyxl import Workbook

from budget_generator.sheets.calculations import register_calculations_named_ranges
from budget_generator.sheets.dashboard import register_dashboard_named_ranges
from budget_generator.sheets.dropdown import register_dropdown_named_ranges
from budget_generator.sheets.planning import build_planning_sheet, register_planning_named_ranges
from budget_generator.sheets.settings import register_settings_named_ranges
from budget_generator.utils.named_ranges import (
    DuplicateNamedRangeError,
    NamedRangeManager,
    NamedRangeSpec,
)


def test_create_range_registers_defined_name() -> None:
    workbook = Workbook()
    manager = NamedRangeManager(workbook)

    manager.create_range("StartingYear", "Settings", "$E$8")

    assert "StartingYear" in workbook.defined_names
    defined = workbook.defined_names["StartingYear"]
    assert defined.attr_text == "Settings!$E$8"


def test_create_constant_registers_literal() -> None:
    workbook = Workbook()
    manager = NamedRangeManager(workbook)

    manager.create_constant("income_min_row", 12)

    defined = workbook.defined_names["income_min_row"]
    assert defined.attr_text == "12"


def test_create_range_quotes_sheet_with_spaces() -> None:
    workbook = Workbook()
    workbook.create_sheet("Dropdown Data")
    manager = NamedRangeManager(workbook)

    manager.create_range("YearsList", "Dropdown Data", "$B$3:$B$7")

    defined = workbook.defined_names["YearsList"]
    assert defined.attr_text == "'Dropdown Data'!$B$3:$B$7"


def test_duplicate_range_raises() -> None:
    workbook = Workbook()
    manager = NamedRangeManager(workbook)
    manager.create_range("StartingYear", "Settings", "$E$8")

    with pytest.raises(DuplicateNamedRangeError):
        manager.create_range("StartingYear", "Settings", "$E$8")


def test_register_many_is_convenience_wrapper() -> None:
    workbook = Workbook()
    specs = [
        NamedRangeSpec("StartingYear", "Settings", "$E$8"),
        NamedRangeSpec("LateIncomeEnabled", "Settings", "$J$16"),
    ]

    manager = NamedRangeManager(workbook)
    manager.register_many(specs)

    assert {
        name for name in workbook.defined_names.keys() if name in {"StartingYear", "LateIncomeEnabled"}
    } == {"StartingYear", "LateIncomeEnabled"}


def test_register_planning_named_ranges_sets_expected_refs() -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Budget-Planning"

    build_planning_sheet(ws, {})

    manager = NamedRangeManager(workbook)
    register_planning_named_ranges(manager)

    income_ref = workbook.defined_names["IncomeCats"].attr_text
    expense_grid_ref = workbook.defined_names["ExpenseGrid"].attr_text
    savings_totals_ref = workbook.defined_names["SavingsTotals"].attr_text
    income_min = workbook.defined_names["income_min_row"].attr_text
    expenses_max = workbook.defined_names["expenses_max_row"].attr_text

    assert income_ref in {"'Budget-Planning'!$D$12:$D$23", "Budget-Planning!$D$12:$D$23"}
    assert expense_grid_ref in {
        "'Budget-Planning'!$E$33:$Q$44",
        "Budget-Planning!$E$33:$Q$44",
    }
    assert savings_totals_ref in {
        "'Budget-Planning'!$E$67:$Q$67",
        "Budget-Planning!$E$67:$Q$67",
    }
    assert income_min == "12"
    assert expenses_max == "44"


def test_register_calculations_named_ranges_sets_expected_refs() -> None:
    workbook = Workbook()
    workbook.create_sheet("Calculations")
    manager = NamedRangeManager(workbook)

    register_calculations_named_ranges(manager)

    month_map_ref = workbook.defined_names["MonthMap"].attr_text
    month_idx_ref = workbook.defined_names["MonthIdx"].attr_text
    assert month_map_ref in {"Calculations!$J$2:$K$13", "'Calculations'!$J$2:$K$13"}
    assert month_idx_ref in {"Calculations!$K$1", "'Calculations'!$K$1"}


def test_register_dashboard_named_ranges_sets_expected_refs() -> None:
    workbook = Workbook()
    workbook.create_sheet("Budget Dashboard")
    manager = NamedRangeManager(workbook)

    register_dashboard_named_ranges(manager)

    dash_year_ref = workbook.defined_names["DashYear"].attr_text
    dash_period_ref = workbook.defined_names["DashPeriod"].attr_text
    assert dash_year_ref in {"'Budget Dashboard'!$C$3", "Budget Dashboard!$C$3"}
    assert dash_period_ref in {"'Budget Dashboard'!$C$4", "Budget Dashboard!$C$4"}


def test_register_settings_named_ranges_sets_expected_refs() -> None:
    workbook = Workbook()
    workbook.create_sheet("Settings")
    manager = NamedRangeManager(workbook)

    register_settings_named_ranges(manager)

    assert workbook.defined_names["StartingYear"].attr_text == "Settings!$E$8"
    assert workbook.defined_names["starting_year"].attr_text == "Settings!$E$8"
    assert workbook.defined_names["LateIncomeEnabled"].attr_text == "Settings!$J$16"


def test_register_dropdown_named_ranges_sets_expected_refs() -> None:
    workbook = Workbook()
    workbook.create_sheet("Dropdown Data")
    manager = NamedRangeManager(workbook)

    register_dropdown_named_ranges(manager)

    years_ref = workbook.defined_names["YearsList"].attr_text
    months_ref = workbook.defined_names["MonthsList"].attr_text
    assert years_ref == "'Dropdown Data'!$B$3:$B$7"
    assert months_ref == "'Dropdown Data'!$C$3:$C$14"
